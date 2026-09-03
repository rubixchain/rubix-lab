#!/usr/bin/env python3
"""
case_runner.py - Generic driver for the master catalogue.

Reuses smoke_test.py's proven common setup (pool -> DID readiness -> role
assignment -> quorum setup+funding -> sender funding) and then runs a
per-asset case module against it. To run a different asset, point --cases at
that module; nothing here changes.

A case module must expose:
    CASES = {"RBT-001": fn, ...}   fn(ctx, ci) -> (passed, actual, note)
    ORDER = ["RBT-001", ...]        execution order

`ci` is a CaseInfo carrying the row from master-test-cases.xlsx (test id,
case text, expected result, other checks, notes) so a case can assert
against what the catalogue actually says rather than a hardcoded copy.

Three outcomes, deliberately distinct:
    PASS    - ran and matched the expectation
    FAIL    - ran and did NOT match: a real finding
    SKIP    - not attempted, with a reason. Never silently counted as a pass.
              Used where the case needs machinery this runner doesn't have
              (node kills, DB seeding, a second DID on one node) or where
              setup would be prohibitively expensive (see the minting note in
              rbt_cases.py). A SKIP is an honest gap, not a success.

Usage:
    python3 case_runner.py --cases rbt            # test-plan/rbt/rbt_cases.py
    python3 case_runner.py --cases rbt --only RBT-001,RBT-005
    python3 case_runner.py --cases rbt --quorum-fund 5000
"""

import argparse
import datetime
import importlib.util
import json
import os
import sys
import time

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import rubix_client as rc
import report_builder
from smoke_test import (
    FIXED_ROLES, DEFAULT_HOSTS, ROLES_PATH,
    sweep_and_prepare, assign_roles, write_roles_file,
    setup_quorums, assign_and_fund_senders,
)

MASTER_PATH = os.path.join(HERE, "master-test-cases.xlsx")

SKIP = "SKIP"


class CaseInfo:
    """One row of master-test-cases.xlsx."""

    def __init__(self, test_id, asset="", case="", expected="", checks="", notes=""):
        self.test_id = test_id
        self.asset = asset
        self.case = case
        self.expected = expected
        self.checks = checks
        self.notes = notes

    @property
    def expects_rejection(self):
        """True when the catalogue's Expected Result is a rejection."""
        return self.expected.strip().lower().startswith("reject")

    @property
    def is_record_only(self):
        """True when the catalogue asks to RECORD behaviour rather than
        assert pass/fail - e.g. value ladders ('Record the largest value that
        works') and 'Define what happens'. Forcing these into pass/fail would
        invent an expectation the catalogue deliberately doesn't state."""
        low = self.expected.strip().lower()
        return low.startswith("record") or low.startswith("define") or "find the" in low


class CaseContext:
    """What every case gets. Same shape as smoke_test's SmokeContext plus
    the extras full-catalogue cases need."""

    def __init__(self, port, quorum_hosts, senders, receivers, sender_quorum, args):
        self.port = port
        self.quorum_hosts = quorum_hosts
        self.senders = senders
        self.receivers = receivers
        self.sender_quorum = sender_quorum
        self.pairs = list(zip(senders, receivers))
        self.args = args

    def pair(self, i=0):
        """A (sender, receiver) pair. Cases that need an isolated pair should
        use different indices so they don't disturb each other's balances."""
        return self.pairs[i % len(self.pairs)]

    def quorum_for(self, sender):
        return self.sender_quorum.get(sender["host"])


def load_master(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("ERROR: openpyxl is required to read the master catalogue.\n"
                 "  sudo apt install -y python3-openpyxl")
    wb = load_workbook(path, read_only=True)
    ws = wb["Master"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        out[row[0]] = CaseInfo(row[0], row[1] or "", row[2] or "",
                                row[3] or "", row[4] or "", row[5] or "")
    return out


def load_case_module(name):
    """Import test-plan/<name>/<name>_cases.py by path."""
    mod_path = os.path.join(HERE, "..", name, "{}_cases.py".format(name))
    mod_path = os.path.abspath(mod_path)
    if not os.path.exists(mod_path):
        sys.exit("ERROR: no case module at {}".format(mod_path))
    spec = importlib.util.spec_from_file_location("{}_cases".format(name), mod_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


class CatalogueReport:
    """Per-Test-ID results, with SKIP tracked separately from PASS/FAIL."""

    def __init__(self):
        self.rows = []

    def record(self, ci, fn, ctx):
        start = time.time()
        try:
            passed, actual, note = fn(ctx, ci)
        except Exception as e:
            passed, actual, note = False, "exception", "{}: {}".format(type(e).__name__, e)
        elapsed = round(time.time() - start, 2)

        # `==` not `is`: SKIP is defined independently in each case module, and
        # Python does not guarantee identity for equal strings across modules.
        if isinstance(passed, str) and passed == SKIP:
            status = "SKIP"
        elif passed is True:
            status = "PASS"
        else:
            status = "FAIL"

        self.rows.append({
            "test_id": ci.test_id, "case": ci.case, "expected": ci.expected,
            "status": status, "actual": actual, "note": note, "seconds": elapsed,
        })
        print("  [{:<4}] {:<9} {:>6.2f}s  {}{}".format(
            status, ci.test_id, elapsed, actual, ("  -- " + note) if note else ""))
        return status

    def save(self, asset, meta, timing_ids):
        paths = rc.new_report_paths("catalogue_{}".format(asset), ("pdf", "json"))
        report_builder.build_json(paths["json"], meta, self.rows)
        pdf_ok = report_builder.build_pdf(
            paths["pdf"], "Rubix Lab - {} Catalogue Run".format(asset.upper()),
            meta, self.rows, timing_ids)

        p = sum(1 for r in self.rows if r["status"] == "PASS")
        f = sum(1 for r in self.rows if r["status"] == "FAIL")
        s = sum(1 for r in self.rows if r["status"] == "SKIP")
        attempted = p + f
        print("\n{} passed, {} failed, {} skipped (of {}).".format(p, f, s, len(self.rows)))
        print("Pass rate of ATTEMPTED cases: {:.1f}% ({}/{}) - skipped are not counted "
              "as passes.".format((100.0 * p / attempted) if attempted else 0.0, p, attempted))
        if f:
            print("FAILED: {}".format(
                ", ".join(r["test_id"] for r in self.rows if r["status"] == "FAIL")))
        if s:
            print("SKIPPED (not attempted, reason in the report's Note column): {}".format(
                ", ".join(r["test_id"] for r in self.rows if r["status"] == "SKIP")))
        print("\nJSON: {}".format(paths["json"]))
        if pdf_ok:
            print("PDF : {}".format(paths["pdf"]))
        else:
            print("PDF : not written - reportlab missing. Results are safe in the JSON "
                  "above. Install with: sudo apt install -y python3-reportlab")


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--cases", required=True, help="asset folder name, e.g. rbt")
    p.add_argument("--only", default="", help="comma-separated Test IDs to run")
    p.add_argument("--hosts", default=DEFAULT_HOSTS)
    p.add_argument("--port", type=int, default=rc.DEFAULT_PORT)
    p.add_argument("--quorum-count", type=int, default=3)
    p.add_argument("--fund-quorum", type=int, default=2000,
                   help="RBT per quorum. A quorum must pledge >= the transfer value "
                        "(core/consensus/checks.go), so this caps testable transfer size.")
    p.add_argument("--fund-sender", type=int, default=200,
                   help="RBT per sender. Minting is one token per unit (~15s per 1000), "
                        "so large values here cost real time.")
    p.add_argument("--rbt-amount", type=float, default=1)
    p.add_argument("--ft-count", type=int, default=10)
    p.add_argument("--ft-token-count", type=int, default=1)
    # Cases whose full catalogue size is impractical in a normal pass. The
    # actual value used is always stated in that case's report row, so a
    # reduced run is never mistaken for the full one.
    p.add_argument("--large-mint", type=int, default=2000,
                   help="RBT-003 single-mint size (~15s per 1000)")
    p.add_argument("--decimal-samples", type=int, default=3,
                   help="RBT-029 values per decimal place (catalogue asks 10)")
    p.add_argument("--repeat-count", type=int, default=25,
                   help="RBT-032 repetitions (catalogue asks 1000, ~33 min)")
    p.add_argument("--version-label", default="not recorded (no version API; "
                                               "get it with controller/node-versions.py)",
                   help="fleet build under test, e.g. '1.0.4' or a branch name. Recorded "
                        "in the report so runs can be compared across releases.")
    args = p.parse_args()

    module = load_case_module(args.cases)
    master = load_master(MASTER_PATH)

    order = list(module.ORDER)
    if args.only:
        wanted = {t.strip() for t in args.only.split(",") if t.strip()}
        order = [t for t in order if t in wanted]
        missing = wanted - set(module.CASES)
        if missing:
            sys.exit("ERROR: unknown Test ID(s): {}".format(", ".join(sorted(missing))))
    if not order:
        sys.exit("ERROR: nothing to run.")

    print("== Common: pool + reachability + DID readiness ==")
    hosts = rc.load_hosts(args.hosts)
    pool = [h for h in hosts if h["role"] not in FIXED_ROLES]
    ready, excluded = sweep_and_prepare(pool, args.port, rc.DEFAULT_TIMEOUT)
    print("{} in pool, {} ready, {} excluded".format(len(pool), len(ready), len(excluded)))
    for line in excluded:
        print("  excluded: {}".format(line))

    print("\n== Common: role assignment ==")
    quorum_hosts, senders, receivers = assign_roles(ready, args.quorum_count)
    write_roles_file(ROLES_PATH, quorum_hosts, senders, receivers)
    print("Quorum: {}  Senders: {}  Receivers: {}".format(
        len(quorum_hosts), len(senders), len(receivers)))

    setup_report = _SetupReport()
    print("\n== Common: quorum setup + funding ==")
    setup_quorums(quorum_hosts, args, setup_report)
    time.sleep(2)
    print("\n== Common: assign quorum + fund senders ==")
    sender_quorum = assign_and_fund_senders(quorum_hosts, senders, args, setup_report)
    if setup_report.failures:
        sys.exit("ERROR: common setup failed on: {}\nFix that before running cases - "
                 "results would be meaningless.".format(", ".join(setup_report.failures)))

    ctx = CaseContext(args.port, quorum_hosts, senders, receivers, sender_quorum, args)

    print("\n== {} cases ({}) ==".format(args.cases.upper(), len(order)))
    time.sleep(3)
    started = time.time()
    started_at = datetime.datetime.now()
    report = CatalogueReport()
    for test_id in order:
        ci = master.get(test_id) or CaseInfo(test_id)
        report.record(ci, module.CASES[test_id], ctx)
    duration = time.time() - started

    # Run conditions - without these a report can't be compared against
    # another run, which is the whole point of a regression lab.
    meta = {
        "asset": args.cases,
        "started_at": started_at.isoformat(timespec="seconds"),
        "duration_seconds": round(duration, 1),
        "conditions": [
            ("Run started", started_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Duration", "{:.0f}s".format(duration)),
            ("Asset / cases run", "{} - {} of {} in the catalogue".format(
                args.cases.upper(), len(order), len(module.ORDER))),
            ("Fleet version", args.version_label),
            ("Hosts file", args.hosts),
            ("Pool hosts ready", "{} (of {} in pool); excluded: {}".format(
                len(ready), len(pool), len(excluded) or "none")),
            ("Quorum hosts", ", ".join("{} ({}...)".format(q["host"], q["did"][:12])
                                        for q in quorum_hosts)),
            ("Sender hosts", ", ".join(s["host"] for s in senders)),
            ("Receiver hosts", ", ".join(r["host"] for r in receivers)),
            ("Sender/receiver pairs", str(len(ctx.pairs))),
            ("Quorum funding", "{} RBT each (caps the largest testable transfer - a "
                               "quorum must pledge >= the transfer value)".format(args.fund_quorum)),
            ("Sender funding", "{} RBT each".format(args.fund_sender)),
            ("Reduced-scale flags", "--large-mint {} | --decimal-samples {} (catalogue asks 10) "
                                     "| --repeat-count {} (catalogue asks 1000)".format(
                                         args.large_mint, args.decimal_samples, args.repeat_count)),
            ("API port", str(args.port)),
        ],
    }
    if args.only:
        meta["conditions"].insert(3, ("Subset filter", "--only {}".format(args.only)))

    timing_ids = set(getattr(module, "TIMING_CASES", set()))
    print()
    report.save(args.cases, meta, timing_ids)


class _SetupReport:
    """Minimal Report-alike for the common setup phase, which predates any
    Test ID. Tracks failures so cases never run against a broken fleet."""

    def __init__(self):
        self.failures = []

    def record(self, step, fn, **params):
        start = time.time()
        try:
            passed, actual, reason = fn()
        except Exception as e:
            passed, actual, reason = False, "exception", "{}: {}".format(type(e).__name__, e)
        elapsed = round(time.time() - start, 2)
        print("  [{:<4}] {:<24} {:>6.2f}s  {}{}".format(
            "PASS" if passed else "FAIL", step, elapsed, actual,
            "" if passed else "  -- " + reason))
        if not passed:
            self.failures.append(step)
        return passed


if __name__ == "__main__":
    main()
