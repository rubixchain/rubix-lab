#!/usr/bin/env python3
"""
run_rbt.py - Pilot RBT test runner. Drives a curated set of "basic
functionality" cases from test-plan/full-test/master-test-cases.xlsx
against the fleet, verifies outcomes via the API, and writes a report.

Case data lives ONLY in master-test-cases.xlsx - this script does not
duplicate the wording, only references Test IDs and executes them.

Requires preflight.py to have run first (reads its context file for which
hosts are quorum/participant-capable and already funded/registered).

Usage:
    cd test-plan/full-test && python3 preflight.py     # once per cycle
    cd ../rbt && python3 run_rbt.py
"""

import datetime
import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "full-test"))
import rubix_client as rc

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl is required. Install it with:\n  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
CONTEXT_PATH = os.path.join(HERE, "..", "full-test", "preflight-context.json")
MASTER_PATH = os.path.join(HERE, "..", "full-test", "master-test-cases.xlsx")
REPORT_PATH = os.path.join(HERE, "rbt-pilot-report.xlsx")

FAKE_BUT_WELLFORMED_DID = "bafybmi" + "z" * 52  # 59 chars, right prefix, never created
BADLY_FORMATTED_DID = "not-a-valid-did"

# Pilot scope: the basic mint/transfer/reject cases only (Setup Method=API
# equivalent). API-RACE / NODE-KILL / MULTI-NODE / DB-SEED cases come later.
PILOT_TEST_IDS = [
    "RBT-001", "RBT-002", "RBT-004", "RBT-005", "RBT-006",
    "RBT-007", "RBT-008", "RBT-024", "RBT-025", "RBT-026",
]


def load_master_lookup(path):
    wb = load_workbook(path, read_only=True)
    ws = wb["Master"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return {row[0]: {"asset": row[1], "case": row[2], "expected": row[3],
                      "checks": row[4], "notes": row[5]} for row in rows}


def load_context(path):
    if not os.path.exists(path):
        sys.exit("ERROR: {} not found - run preflight.py first.".format(path))
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


def close_enough(a, b, tol=0.0015):
    return abs(a - b) <= tol


def run_case(test_id, fn, results):
    """fn() -> (passed: bool, actual: str, note: str)"""
    start = datetime.datetime.now()
    try:
        passed, actual, note = fn()
    except Exception as e:
        passed, actual, note = False, "exception", "{}: {}".format(type(e).__name__, e)
    elapsed = (datetime.datetime.now() - start).total_seconds()
    results.append({"test_id": test_id, "passed": passed, "actual": actual,
                     "note": note, "seconds": round(elapsed, 2)})
    print("  {:<10} {:<6} {:<8.2f}s  {}  {}".format(
        test_id, "PASS" if passed else "FAIL", elapsed, actual, note))


def main():
    print("== RBT pilot run ==")
    ctx = load_context(CONTEXT_PATH)
    port = ctx["port"]
    participants = ctx["participant_hosts"]
    if len(participants) < 2:
        sys.exit("ERROR: need at least 2 participant hosts in the preflight context.")
    a, b = participants[0], participants[1]

    lookup = load_master_lookup(MASTER_PATH)
    results = []

    print("Using A={} B={}\n".format(a["host"], b["host"]))

    # --- RBT-001: Mint local RBT on a node ---
    def case_mint():
        ok0, bal0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        status, message = rc.fund_did(a["host"], a["did"], 50, port)
        if not status:
            return False, "mint rejected", message
        confirmed, bal1 = rc.wait_for_balance(a["host"], a["did"], (bal0 or 0) + 50, port)
        if not confirmed:
            return False, "balance did not rise as expected", "before={} after={}".format(bal0, bal1)
        return True, "balance {} -> {}".format(bal0, bal1), ""
    run_case("RBT-001", case_mint, results)

    # --- RBT-002: Mint RBT twice on the same node ---
    def case_mint_twice():
        ok0, bal0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        s1, m1 = rc.fund_did(a["host"], a["did"], 25, port)
        s2, m2 = rc.fund_did(a["host"], a["did"], 25, port)
        if not (s1 and s2):
            return False, "one or both mints rejected", "{} / {}".format(m1, m2)
        confirmed, bal1 = rc.wait_for_balance(a["host"], a["did"], (bal0 or 0) + 50, port)
        if not confirmed:
            return False, "combined balance did not add up", "before={} after={}".format(bal0, bal1)
        return True, "balance {} -> {} (two mints of 25)".format(bal0, bal1), ""
    run_case("RBT-002", case_mint_twice, results)

    # --- RBT-004: Mint on a node that has no DID yet ---
    def case_mint_no_did():
        return False, "SKIPPED", ("no no-DID host available in the current pool - every ready host "
                                   "already has a DID from earlier dids-to-excel.py runs. Needs a "
                                   "freshly-installed, DID-less node to test properly.")
    run_case("RBT-004", case_mint_no_did, results)

    # --- RBT-005: Send 1 RBT from A to B ---
    def case_send_a_to_b():
        ok_a0, bal_a0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        ok_b0, bal_b0, _ = rc.get_rbt_balance(b["host"], b["did"], port)
        status, message, _ = rc.initiate_transaction(a["host"], a["did"], a["did"],
                                                       rbt=1, memo="RBT-005 pilot", port=port)
        if not status:
            return False, "transfer rejected", message
        ok_a1, bal_a1, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        ok_b1, bal_b1, _ = rc.get_rbt_balance(b["host"], b["did"], port)
        if close_enough(bal_a1, bal_a0 - 1) and close_enough(bal_b1, bal_b0 + 1):
            return True, "A {}->{}  B {}->{}".format(bal_a0, bal_a1, bal_b0, bal_b1), ""
        return False, "balances did not move by exactly 1", "A {}->{} B {}->{}".format(
            bal_a0, bal_a1, bal_b0, bal_b1)
    run_case("RBT-005", case_send_a_to_b, results)

    # --- RBT-006: Send RBT back from B to A ---
    def case_send_b_to_a():
        ok_a0, bal_a0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        ok_b0, bal_b0, _ = rc.get_rbt_balance(b["host"], b["did"], port)
        status, message, _ = rc.initiate_transaction(b["host"], b["did"], b["did"],
                                                       rbt=1, memo="RBT-006 pilot", port=port)
        if not status:
            return False, "transfer rejected", message
        ok_a1, bal_a1, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        ok_b1, bal_b1, _ = rc.get_rbt_balance(b["host"], b["did"], port)
        if close_enough(bal_b1, bal_b0 - 1) and close_enough(bal_a1, bal_a0 + 1):
            return True, "B {}->{}  A {}->{}".format(bal_b0, bal_b1, bal_a0, bal_a1), ""
        return False, "balances did not move by exactly 1", "A {}->{} B {}->{}".format(
            bal_a0, bal_a1, bal_b0, bal_b1)
    run_case("RBT-006", case_send_b_to_a, results)

    # --- RBT-007: Send RBT to a DID that does not exist ---
    def case_send_nonexistent_did():
        ok0, bal0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        status, message, _ = rc.initiate_transaction(a["host"], FAKE_BUT_WELLFORMED_DID, a["did"],
                                                       rbt=1, memo="RBT-007 pilot", port=port)
        ok1, bal1, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        if status:
            return False, "transfer to a nonexistent DID SUCCEEDED (should reject)", message
        if not close_enough(bal0, bal1):
            return False, "rejected as expected, but balance changed", "before={} after={}".format(bal0, bal1)
        return True, "rejected: {}".format(message), ""
    run_case("RBT-007", case_send_nonexistent_did, results)

    # --- RBT-008: Send RBT to a badly formatted DID ---
    def case_send_badly_formatted_did():
        ok0, bal0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        status, message, _ = rc.initiate_transaction(a["host"], BADLY_FORMATTED_DID, a["did"],
                                                       rbt=1, memo="RBT-008 pilot", port=port)
        ok1, bal1, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        if status:
            return False, "transfer to a badly formatted DID SUCCEEDED (should reject)", message
        if not close_enough(bal0, bal1):
            return False, "rejected as expected, but balance changed", "before={} after={}".format(bal0, bal1)
        return True, "rejected: {}".format(message), ""
    run_case("RBT-008", case_send_badly_formatted_did, results)

    # --- RBT-024: Send 0 RBT ---
    def case_send_zero():
        ok0, bal0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        status, message, _ = rc.initiate_transaction(a["host"], b["did"], a["did"],
                                                       rbt=0, memo="RBT-024 pilot", port=port)
        ok1, bal1, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        if status:
            return False, "sending 0 RBT SUCCEEDED (should reject)", message
        if not close_enough(bal0, bal1):
            return False, "rejected as expected, but balance changed", "before={} after={}".format(bal0, bal1)
        return True, "rejected: {}".format(message), ""
    run_case("RBT-024", case_send_zero, results)

    # --- RBT-025: Send a negative amount ---
    def case_send_negative():
        ok0, bal0, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        status, message, _ = rc.initiate_transaction(a["host"], b["did"], a["did"],
                                                       rbt=-1, memo="RBT-025 pilot", port=port)
        ok1, bal1, _ = rc.get_rbt_balance(a["host"], a["did"], port)
        if status:
            return False, "sending a negative amount SUCCEEDED (should reject)", message
        if not close_enough(bal0, bal1):
            return False, "rejected as expected, but balance changed", "before={} after={}".format(bal0, bal1)
        return True, "rejected: {}".format(message), ""
    run_case("RBT-025", case_send_negative, results)

    # --- RBT-026: Send more than the wallet holds ---
    def case_send_more_than_held():
        ok0, bal0, _ = rc.get_rbt_balance(b["host"], b["did"], port)
        huge = (bal0 or 0) + 1_000_000
        status, message, _ = rc.initiate_transaction(b["host"], a["did"], b["did"],
                                                       rbt=huge, memo="RBT-026 pilot", port=port)
        ok1, bal1, _ = rc.get_rbt_balance(b["host"], b["did"], port)
        if status:
            return False, "sending more than the wallet holds SUCCEEDED (should reject)", message
        if not close_enough(bal0, bal1):
            return False, "rejected as expected, but balance changed", "before={} after={}".format(bal0, bal1)
        return True, "rejected: {}".format(message), ""
    run_case("RBT-026", case_send_more_than_held, results)

    # --- Report ---
    wb = Workbook()
    ws = wb.active
    ws.title = "RBT Pilot"
    ws.append(["Test ID", "Test Case", "Expected Result", "Pass/Fail", "Actual",
               "Note", "Seconds", "Run At"])
    run_at = datetime.datetime.now().isoformat(timespec="seconds")
    for r in results:
        m = lookup.get(r["test_id"], {})
        ws.append([r["test_id"], m.get("case", ""), m.get("expected", ""),
                   "PASS" if r["passed"] else "FAIL", r["actual"], r["note"],
                   r["seconds"], run_at])
    for col, width in zip("ABCDEFGH", (10, 45, 30, 10, 40, 45, 8, 20)):
        ws.column_dimensions[col].width = width
    wb.save(REPORT_PATH)

    passed = sum(1 for r in results if r["passed"])
    print("\n{}/{} passed. Report: {}".format(passed, len(results), REPORT_PATH))


if __name__ == "__main__":
    main()
